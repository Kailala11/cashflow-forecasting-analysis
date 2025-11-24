"""
===================================================================================
CASH FLOW FORECASTING ANALYSIS - CAFÉ SUMARECON BEKASI
===================================================================================

Portfolio Project: Financial Analysis & Data Visualization

Author: Data Analyst Portfolio
Date: November 2024
Tools: Python, Pandas, Matplotlib, Seaborn, Openpyxl

This script performs comprehensive financial analysis on cash flow data including:
- Multi-scenario comparison
- Statistical analysis (volatility, correlation, trends)
- Break-even analysis
- Data visualization
- Risk assessment

Requirements:
    pip install pandas numpy matplotlib seaborn openpyxl

Usage:
    python cashflow_analysis.py
    
Input:
    cafe_cashflow_bekasi.xlsx (should be in same directory)
    
Output:
    - PNG visualizations
    - CSV summary files
    - Console analysis report
===================================================================================
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from datetime import datetime
import warnings
import os

warnings.filterwarnings('ignore')

# Set style for professional visualizations
plt.style.use('seaborn-v0_8-darkgrid')
sns.set_palette("husl")

print("="*80)
print("CASH FLOW FORECASTING ANALYSIS - CAFÉ SUMARECON BEKASI")
print("="*80)
print(f"\nAnalysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("\n" + "="*80)

# Check if file exists
excel_file = 'cafe_cashflow_bekasi.xlsx'
if not os.path.exists(excel_file):
    print(f"❌ ERROR: File '{excel_file}' not found!")
    print("   Please ensure the Excel file is in the same directory as this script.")
    exit(1)

print(f"✓ Loading data from: {excel_file}")

# ============= DATA EXTRACTION =============
print("\n📊 EXTRACTING DATA FROM MULTIPLE SCENARIOS...")
print("-" * 80)

# Load the workbook
wb = load_workbook(excel_file, data_only=True)

def extract_scenario_data(sheet_name):
    """Extract cash flow data from a scenario sheet"""
    ws = wb[sheet_name]
    
    months = []
    opening = []
    inflows = []
    outflows = []
    net_cf = []
    closing = []
    
    # Get month headers (row 4, columns B-J)
    for col in range(2, 11):
        month_val = ws.cell(4, col).value
        months.append(month_val if month_val else f"Month {col-1}")
    
    # Opening balance (row 5)
    for col in range(2, 11):
        opening.append(ws.cell(5, col).value or 0)
    
    # Total inflows (row 13)
    for col in range(2, 11):
        inflows.append(ws.cell(13, col).value or 0)
    
    # Total outflows (row 25)
    for col in range(2, 11):
        outflows.append(ws.cell(25, col).value or 0)
    
    # Net cash flow (row 27)
    for col in range(2, 11):
        net_cf.append(ws.cell(27, col).value or 0)
    
    # Closing balance (row 29)
    for col in range(2, 11):
        closing.append(ws.cell(29, col).value or 0)
    
    df = pd.DataFrame({
        'Month': months,
        'Opening_Balance': opening,
        'Total_Inflows': inflows,
        'Total_Outflows': outflows,
        'Net_Cash_Flow': net_cf,
        'Closing_Balance': closing
    })
    
    return df

# Extract all scenarios
try:
    base_df = extract_scenario_data('Skenario Base')
    optimistic_df = extract_scenario_data('Skenario Optimistis')
    pessimistic_df = extract_scenario_data('Skenario Pesimistis')
    
    print(f"✓ Data extracted successfully!")
    print(f"  - Periods analyzed: {len(base_df)} months")
    print(f"  - Scenarios: 3 (Base, Optimistic, Pessimistic)")
    print(f"  - Date range: {base_df['Month'].iloc[0]} to {base_df['Month'].iloc[-1]}")
except Exception as e:
    print(f"❌ Error extracting data: {e}")
    exit(1)

# ============= DESCRIPTIVE STATISTICS =============
print("\n" + "="*80)
print("📈 DESCRIPTIVE STATISTICS - BASE CASE SCENARIO")
print("="*80)

stats_cols = ['Total_Inflows', 'Total_Outflows', 'Net_Cash_Flow', 'Closing_Balance']
stats_summary = base_df[stats_cols].describe()

print("\nSummary Statistics (in Million IDR):")
print((stats_summary / 1_000_000).round(2))

# Calculate key metrics
avg_inflows = base_df['Total_Inflows'].mean()
avg_outflows = base_df['Total_Outflows'].mean()
avg_net_cf = base_df['Net_Cash_Flow'].mean()
total_inflows = base_df['Total_Inflows'].sum()
total_outflows = base_df['Total_Outflows'].sum()

# Growth rate (comparing first and last month)
if base_df['Total_Inflows'].iloc[0] > 0:
    growth_rate = ((base_df['Total_Inflows'].iloc[-1] / base_df['Total_Inflows'].iloc[0]) - 1) * 100
else:
    growth_rate = 0

print(f"\n📊 Key Metrics (Base Case):")
print(f"  • Average Monthly Inflows:  Rp {avg_inflows/1_000_000:>10,.1f} Juta")
print(f"  • Average Monthly Outflows: Rp {avg_outflows/1_000_000:>10,.1f} Juta")
print(f"  • Average Net Cash Flow:    Rp {avg_net_cf/1_000_000:>10,.1f} Juta")
print(f"  • Total Inflows (Period):   Rp {total_inflows/1_000_000:>10,.1f} Juta")
print(f"  • Total Outflows (Period):  Rp {total_outflows/1_000_000:>10,.1f} Juta")
print(f"  • Revenue Growth Rate:      {growth_rate:>10.1f}%")
print(f"  • Expense Ratio:            {(avg_outflows/avg_inflows*100):>10.1f}%")

# ============= SCENARIO COMPARISON =============
print("\n" + "="*80)
print("🔍 SCENARIO COMPARISON ANALYSIS")
print("="*80)

scenarios_comparison = pd.DataFrame({
    'Metric': [
        'Final Cash Balance', 
        'Total Cash Inflow', 
        'Total Cash Outflow', 
        'Cumulative Net CF',
        'Average Monthly CF'
    ],
    'Optimistic (Rp M)': [
        optimistic_df['Closing_Balance'].iloc[-1] / 1_000_000,
        optimistic_df['Total_Inflows'].sum() / 1_000_000,
        optimistic_df['Total_Outflows'].sum() / 1_000_000,
        optimistic_df['Net_Cash_Flow'].sum() / 1_000_000,
        optimistic_df['Net_Cash_Flow'].mean() / 1_000_000
    ],
    'Base Case (Rp M)': [
        base_df['Closing_Balance'].iloc[-1] / 1_000_000,
        base_df['Total_Inflows'].sum() / 1_000_000,
        base_df['Total_Outflows'].sum() / 1_000_000,
        base_df['Net_Cash_Flow'].sum() / 1_000_000,
        base_df['Net_Cash_Flow'].mean() / 1_000_000
    ],
    'Pessimistic (Rp M)': [
        pessimistic_df['Closing_Balance'].iloc[-1] / 1_000_000,
        pessimistic_df['Total_Inflows'].sum() / 1_000_000,
        pessimistic_df['Total_Outflows'].sum() / 1_000_000,
        pessimistic_df['Net_Cash_Flow'].sum() / 1_000_000,
        pessimistic_df['Net_Cash_Flow'].mean() / 1_000_000
    ]
})

print("\nScenario Comparison:")
print(scenarios_comparison.to_string(index=False))

# Calculate risk metrics
best_case = optimistic_df['Closing_Balance'].iloc[-1]
worst_case = pessimistic_df['Closing_Balance'].iloc[-1]
base_case = base_df['Closing_Balance'].iloc[-1]
risk_range = best_case - worst_case

print(f"\n⚠️ Risk Analysis:")
print(f"  • Best Case Scenario:  Rp {best_case/1_000_000:>10,.1f} Juta")
print(f"  • Base Case Scenario:  Rp {base_case/1_000_000:>10,.1f} Juta")
print(f"  • Worst Case Scenario: Rp {worst_case/1_000_000:>10,.1f} Juta")
print(f"  • Risk Range:          Rp {risk_range/1_000_000:>10,.1f} Juta")
print(f"  • Scenario Spread:     {(risk_range/base_case*100):>10.1f}%")

# ============= STATISTICAL ANALYSIS =============
print("\n" + "="*80)
print("📐 STATISTICAL ANALYSIS")
print("="*80)

# Calculate volatility (standard deviation)
base_volatility = base_df['Net_Cash_Flow'].std() / 1_000_000
opt_volatility = optimistic_df['Net_Cash_Flow'].std() / 1_000_000
pes_volatility = pessimistic_df['Net_Cash_Flow'].std() / 1_000_000

print(f"\n📊 Cash Flow Volatility (Standard Deviation):")
print(f"  • Base Case:    Rp {base_volatility:>8,.1f} Juta")
print(f"  • Optimistic:   Rp {opt_volatility:>8,.1f} Juta")
print(f"  • Pessimistic:  Rp {pes_volatility:>8,.1f} Juta")

# Calculate correlation between inflows and outflows
if base_df['Total_Inflows'].std() > 0 and base_df['Total_Outflows'].std() > 0:
    correlation = base_df['Total_Inflows'].corr(base_df['Total_Outflows'])
    print(f"\n🔗 Correlation (Inflows vs Outflows): {correlation:.3f}")
    if correlation > 0.8:
        print("  → Strong positive correlation detected")
    elif correlation > 0.5:
        print("  → Moderate positive correlation")
    else:
        print("  → Weak correlation")
else:
    print(f"\n🔗 Correlation: Cannot calculate (insufficient variance)")

# Trend analysis
months_numeric = np.arange(len(base_df))
if len(months_numeric) > 1:
    inflow_trend = np.polyfit(months_numeric, base_df['Total_Inflows'], 1)
    outflow_trend = np.polyfit(months_numeric, base_df['Total_Outflows'], 1)
    
    print(f"\n📈 Trend Analysis (Base Case):")
    print(f"  • Inflow trend:  Rp {inflow_trend[0]/1_000_000:>8,.2f} M per month")
    print(f"  • Outflow trend: Rp {outflow_trend[0]/1_000_000:>8,.2f} M per month")

# ============= BREAK-EVEN ANALYSIS =============
print("\n" + "="*80)
print("⚖️ BREAK-EVEN ANALYSIS")
print("="*80)

try:
    ws_be = wb['Analisis Break-Even']
    be_revenue = ws_be.cell(11, 2).value or 0
    be_transactions = ws_be.cell(10, 2).value or 0
    current_revenue = ws_be.cell(14, 2).value or 0
    safety_margin = ws_be.cell(18, 2).value or 0
    
    print(f"\n💰 Break-Even Metrics:")
    print(f"  • BE Revenue:        Rp {be_revenue/1_000_000:>10,.1f} Juta/bulan")
    print(f"  • BE Transactions:   {be_transactions:>10,.0f} transaksi/bulan")
    print(f"  • Current Revenue:   Rp {current_revenue/1_000_000:>10,.1f} Juta/bulan")
    print(f"  • Safety Margin:     {safety_margin*100:>10.1f}%")
    
    if safety_margin > 0.15:
        status = "AMAN (margin keamanan tinggi)"
        icon = "🟢"
    elif safety_margin > 0.08:
        status = "CUKUP AMAN (perlu monitoring)"
        icon = "🟡"
    else:
        status = "PERLU PERHATIAN (margin tipis)"
        icon = "🔴"
    
    print(f"  {icon} Status: {status}")
except Exception as e:
    print(f"⚠️ Could not extract break-even data: {e}")

# ============= VISUALIZATIONS =============
print("\n" + "="*80)
print("📊 GENERATING VISUALIZATIONS")
print("="*80)

try:
    # Create comprehensive dashboard
    fig, axes = plt.subplots(2, 2, figsize=(14, 10))
    fig.suptitle('Cash Flow Forecasting Analysis - Café Sumarecon Bekasi', 
                 fontsize=16, fontweight='bold', y=0.995)
    
    # Chart 1: Multi-Scenario Cash Balance
    ax1 = axes[0, 0]
    ax1.plot(base_df['Month'], base_df['Closing_Balance']/1_000_000_000, 
             marker='o', linewidth=2.5, label='Base Case', color='#2E86AB')
    ax1.plot(optimistic_df['Month'], optimistic_df['Closing_Balance']/1_000_000_000, 
             marker='s', linewidth=2.5, label='Optimistic', color='#06A77D')
    ax1.plot(pessimistic_df['Month'], pessimistic_df['Closing_Balance']/1_000_000_000, 
             marker='^', linewidth=2.5, label='Pessimistic', color='#D72638')
    ax1.set_title('Cash Balance Projection', fontsize=13, fontweight='bold')
    ax1.set_ylabel('Balance (Billion IDR)', fontsize=10)
    ax1.legend(loc='best', fontsize=9)
    ax1.grid(True, alpha=0.3)
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right', fontsize=8)
    
    # Chart 2: Final Balance Comparison
    ax2 = axes[0, 1]
    scenarios = ['Optimistic', 'Base Case', 'Pessimistic']
    final_vals = [optimistic_df['Closing_Balance'].iloc[-1]/1_000_000_000,
                  base_df['Closing_Balance'].iloc[-1]/1_000_000_000,
                  pessimistic_df['Closing_Balance'].iloc[-1]/1_000_000_000]
    colors = ['#06A77D', '#2E86AB', '#D72638']
    bars = ax2.barh(scenarios, final_vals, color=colors, alpha=0.8)
    ax2.set_title('Final Balance Comparison', fontsize=13, fontweight='bold')
    ax2.set_xlabel('Balance (Billion IDR)', fontsize=10)
    for bar, val in zip(bars, final_vals):
        ax2.text(val + 0.03, bar.get_y() + bar.get_height()/2, 
                f'Rp {val:.2f}B', va='center', fontsize=9)
    ax2.grid(True, axis='x', alpha=0.3)
    
    # Chart 3: Month-over-Month Growth
    ax3 = axes[1, 0]
    base_growth = []
    for i in range(len(base_df)):
        if i > 0 and base_df['Closing_Balance'].iloc[i-1] > 0:
            growth = ((base_df['Closing_Balance'].iloc[i] / 
                      base_df['Closing_Balance'].iloc[i-1]) - 1) * 100
        else:
            growth = 0
        base_growth.append(growth)
    
    ax3.plot(base_df['Month'], base_growth, marker='o', linewidth=2, color='#2E86AB')
    ax3.axhline(y=0, color='black', linestyle='-', linewidth=0.8)
    ax3.set_title('Month-over-Month Growth Rate', fontsize=13, fontweight='bold')
    ax3.set_ylabel('Growth (%)', fontsize=10)
    ax3.grid(True, alpha=0.3)
    plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45, ha='right', fontsize=8)
    
    # Chart 4: Cumulative Cash Flow
    ax4 = axes[1, 1]
    base_cumulative = base_df['Net_Cash_Flow'].cumsum() / 1_000_000_000
    opt_cumulative = optimistic_df['Net_Cash_Flow'].cumsum() / 1_000_000_000
    pes_cumulative = pessimistic_df['Net_Cash_Flow'].cumsum() / 1_000_000_000
    
    x_range = range(len(base_df))
    ax4.fill_between(x_range, opt_cumulative, pes_cumulative, 
                     alpha=0.2, color='gray', label='Uncertainty Range')
    ax4.plot(base_cumulative, marker='o', linewidth=2.5, label='Base Case', color='#2E86AB')
    ax4.plot(opt_cumulative, linestyle='--', linewidth=2, label='Optimistic', color='#06A77D')
    ax4.plot(pes_cumulative, linestyle='--', linewidth=2, label='Pessimistic', color='#D72638')
    ax4.set_title('Cumulative Cash Flow', fontsize=13, fontweight='bold')
    ax4.set_xlabel('Period', fontsize=10)
    ax4.set_ylabel('Cumulative CF (Billion IDR)', fontsize=10)
    ax4.legend(fontsize=8, loc='best')
    ax4.grid(True, alpha=0.3)
    
    plt.tight_layout()
    output_file = 'cashflow_analysis_dashboard.png'
    plt.savefig(output_file, dpi=200, bbox_inches='tight')
    plt.close()
    print(f"✓ Dashboard saved: {output_file}")
    
except Exception as e:
    print(f"⚠️ Error generating visualizations: {e}")

# ============= EXPORT SUMMARY =============
print("\n" + "="*80)
print("💾 EXPORTING ANALYSIS RESULTS")
print("="*80)

try:
    # Create summary dataframe
    summary_df = pd.DataFrame({
        'Metric': [
            'Avg Monthly Revenue',
            'Avg Monthly Expenses',
            'Avg Net Cash Flow',
            'Revenue Growth Rate',
            'Expense Ratio',
            'Cash Flow Volatility',
            'Final Balance (Base)',
            'Final Balance (Best)',
            'Final Balance (Worst)',
            'Risk Range'
        ],
        'Value': [
            f"Rp {avg_inflows/1_000_000:,.1f}M",
            f"Rp {avg_outflows/1_000_000:,.1f}M",
            f"Rp {avg_net_cf/1_000_000:,.1f}M",
            f"{growth_rate:.1f}%",
            f"{(avg_outflows/avg_inflows*100):.1f}%",
            f"Rp {base_volatility:.1f}M",
            f"Rp {base_case/1_000_000:,.1f}M",
            f"Rp {best_case/1_000_000:,.1f}M",
            f"Rp {worst_case/1_000_000:,.1f}M",
            f"Rp {risk_range/1_000_000:,.1f}M"
        ]
    })
    
    summary_file = 'analysis_summary.csv'
    summary_df.to_csv(summary_file, index=False)
    print(f"✓ Summary exported: {summary_file}")
    
    # Export detailed scenario data
    combined_df = pd.DataFrame({
        'Month': base_df['Month'],
        'Base_Inflows': base_df['Total_Inflows'],
        'Base_Outflows': base_df['Total_Outflows'],
        'Base_NetCF': base_df['Net_Cash_Flow'],
        'Base_Balance': base_df['Closing_Balance'],
        'Opt_Balance': optimistic_df['Closing_Balance'],
        'Pes_Balance': pessimistic_df['Closing_Balance']
    })
    
    scenario_file = 'scenario_comparison.csv'
    combined_df.to_csv(scenario_file, index=False)
    print(f"✓ Scenario data exported: {scenario_file}")
    
except Exception as e:
    print(f"⚠️ Error exporting data: {e}")

# ============= COMPLETION =============
print("\n" + "="*80)
print("✅ ANALYSIS COMPLETE!")
print("="*80)
print("\n📁 Generated Files:")
print("  1. cashflow_analysis_dashboard.png - Main visualization dashboard")
print("  2. analysis_summary.csv - Key metrics summary")
print("  3. scenario_comparison.csv - Detailed scenario data")
print("\n" + "="*80)
print(f"Analysis completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("="*80 + "\n")
